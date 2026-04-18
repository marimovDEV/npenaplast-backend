import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal
from django.db.models import Sum, Avg
from sales_v2.models import Invoice, SaleItem, Customer
from production_v2.models import ProductionBatch
from finance_v2.models import ClientBalance

def generate_enterprise_xlsx(report_type, start_date=None, end_date=None):
    """
    Generates high-premium Excel reports for Director/Accountant (Phase 7).
    Types: 'PROFIT_LOSS', 'PRODUCTION', 'DEBT_AGING', 'INVENTORY'
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800
    border = Border(left=Side(style='thin', color="CBD5E1"), 
                    right=Side(style='thin', color="CBD5E1"),
                    top=Side(style='thin', color="CBD5E1"), 
                    bottom=Side(style='thin', color="CBD5E1"))

    if report_type == 'PROFIT_LOSS':
        _fill_pl_report(ws, start_date, end_date, header_font, header_fill, border)
    elif report_type == 'PRODUCTION':
        _fill_production_report(ws, start_date, end_date, header_font, header_fill, border)
    elif report_type == 'DEBT_AGING':
        _fill_debt_report(ws, header_font, header_fill, border)

    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 5

    return wb

def _fill_pl_report(ws, start_date, end_date, h_font, h_fill, border):
    ws.append(["Yuksar ERP+ - Foyda va Zarar Hisoboti (P&L)"])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=16)
    ws.append([f"Davr: {start_date or 'Barcha'} - {end_date or 'Bugun'}"])
    ws.append([]) # Spacer

    headers = ["Kategoriya", "Tavsif", "Summa (UZS)", "Ulush %"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = h_font
        cell.fill = h_fill
        cell.border = border

    # Data Calculation
    invoices = Invoice.objects.filter(status='COMPLETED')
    if start_date: invoices = invoices.filter(date__gte=start_date)
    if end_date: invoices = invoices.filter(date__lte=end_date)
    
    total_rev = invoices.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')
    
    # Simple P&L Logic
    data = [
        ["DAROMAD", "Jami sotuvlar", float(total_rev), "100%"],
        ["XARAJAT", "Xom-ashyo tannarxi (Taxminiy)", float(total_rev * Decimal('0.6')), "60%"],
        ["XARAJAT", "Ish haqi", float(total_rev * Decimal('0.1')), "10%"],
        ["XARAJAT", "Overhead / Elektr", float(total_rev * Decimal('0.05')), "5%"],
        [],
        ["OPERATSION FOYDA", "", float(total_rev * Decimal('0.25')), "25%"]
    ]

    for row in data:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            if row and row[0] in ["DAROMAD", "OPERATSION FOYDA"]:
                cell.font = Font(bold=True)

def _fill_production_report(ws, start_date, end_date, h_font, h_fill, border):
    headers = ["Batch No", "Mahsulot", "Hajm (m3)", "Tannarx (Unit)", "Jami Tannarx", "Sana"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = h_font
        cell.fill = h_fill
        cell.border = border

    batches = ProductionBatch.objects.all().order_by('-created_at')
    for b in batches:
        ws.append([
            b.batch_number,
            b.product.name if b.product else "N/A",
            float(b.total_volume),
            float(b.unit_cost),
            float(b.total_cost),
            b.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        for cell in ws[ws.max_row]:
            cell.border = border

def _fill_debt_report(ws, h_font, h_fill, border):
    ws.append(["Muddati o'tgan qarzdorlik - Komplayens Hisoboti"])
    headers = ["Mijoz", "Telefon", "Jami Qarz", "Muddati O'tgan", "Holat"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = h_font
        cell.fill = h_fill
        cell.border = border

    balances = ClientBalance.objects.all().order_by('-overdue_debt')
    for cb in balances:
        ws.append([
            cb.customer.name,
            cb.customer.phone,
            float(cb.total_debt),
            float(cb.overdue_debt),
            "RISK" if cb.overdue_debt > 0 else "HEALTHY"
        ])
        for cell in ws[ws.max_row]:
            cell.border = border
            if cb.overdue_debt > 0:
                cell.style = 'Bad' # Excel built-in style or manual red
